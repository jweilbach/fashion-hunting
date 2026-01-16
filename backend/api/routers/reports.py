"""
Reports Router
Handles CRUD operations for media reports
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
from uuid import UUID
import sys
from pathlib import Path
from math import ceil
import csv
import io

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "src"))

from api.database import get_db
from api import schemas
from api.auth import get_current_user, require_viewer, require_editor
from models.user import User
from repositories.report_repository import ReportRepository

router = APIRouter()


@router.get("/", response_model=schemas.ReportListResponse)
async def list_reports(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=100, description="Items per page"),
    provider: Optional[str] = Query(None, description="Filter by provider (RSS, TikTok, etc)"),
    sentiment: Optional[str] = Query(None, description="Filter by sentiment (positive, neutral, negative)"),
    brand: Optional[str] = Query(None, description="Filter by brand name"),
    status: Optional[str] = Query(None, description="Filter by processing status"),
    start_date: Optional[datetime] = Query(None, description="Filter from date"),
    end_date: Optional[datetime] = Query(None, description="Filter to date"),
    search: Optional[str] = Query(None, description="Search in title and summary"),
    source_type: Optional[str] = Query(None, description="Filter by source type (social, digital)"),
    current_user: User = Depends(require_viewer),
    db: Session = Depends(get_db)
):
    """
    List all reports for the current user's tenant with filtering and pagination

    - **page**: Page number (default: 1)
    - **page_size**: Items per page (default: 50, max: 100)
    - **provider**: Filter by provider
    - **sentiment**: Filter by sentiment (positive, neutral, negative)
    - **brand**: Filter by brand name
    - **status**: Filter by processing status
    - **start_date**: Filter from date
    - **end_date**: Filter to date
    - **search**: Search in title and summary
    - **source_type**: Filter by source type (social, digital)
    """
    repo = ReportRepository(db)

    # Calculate offset
    skip = (page - 1) * page_size

    # Get reports with filters
    if search:
        reports = repo.search(current_user.tenant_id, search, limit=page_size)
        # Count total for search (approximate)
        total = len(reports)
    else:
        reports = repo.get_all(
            tenant_id=current_user.tenant_id,
            skip=skip,
            limit=page_size,
            provider=provider,
            status=status or 'completed',
            start_date=start_date,
            end_date=end_date,
            sentiment=sentiment,
            brand=brand,
            source_type=source_type
        )

        # Get total count efficiently using count method
        total = repo.count(
            tenant_id=current_user.tenant_id,
            provider=provider,
            status=status or 'completed',
            start_date=start_date,
            end_date=end_date,
            sentiment=sentiment,
            brand=brand,
            source_type=source_type
        )

    # Calculate pages
    pages = ceil(total / page_size)

    return schemas.ReportListResponse(
        items=[schemas.Report.model_validate(r) for r in reports],
        total=total,
        page=page,
        page_size=page_size,
        pages=pages
    )


@router.get("/{report_id}", response_model=schemas.Report)
async def get_report(
    report_id: UUID,
    current_user: User = Depends(require_viewer),
    db: Session = Depends(get_db)
):
    """Get a specific report by ID"""
    repo = ReportRepository(db)
    report = repo.get_by_id(report_id)

    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report not found"
        )

    # Verify tenant access
    if report.tenant_id != current_user.tenant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied"
        )

    return schemas.Report.model_validate(report)


@router.get("/brand/{brand_name}", response_model=List[schemas.Report])
async def get_reports_by_brand(
    brand_name: str,
    limit: int = Query(100, ge=1, le=500),
    current_user: User = Depends(require_viewer),
    db: Session = Depends(get_db)
):
    """Get all reports mentioning a specific brand"""
    repo = ReportRepository(db)
    reports = repo.get_by_brand(current_user.tenant_id, brand_name, limit=limit)

    return [schemas.Report.model_validate(r) for r in reports]


@router.patch("/{report_id}", response_model=schemas.Report)
@router.put("/{report_id}", response_model=schemas.Report)
async def update_report(
    report_id: UUID,
    report_update: schemas.ReportUpdate,
    current_user: User = Depends(require_viewer),
    db: Session = Depends(get_db)
):
    """Update a report (sentiment, topic, brands, status)"""
    repo = ReportRepository(db)
    report = repo.get_by_id(report_id)

    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report not found"
        )

    # Verify tenant access
    if report.tenant_id != current_user.tenant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied"
        )

    # Update fields
    update_data = report_update.model_dump(exclude_unset=True)
    updated_report = repo.update(report_id, **update_data)

    return schemas.Report.model_validate(updated_report)


@router.delete("/{report_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_report(
    report_id: UUID,
    current_user: User = Depends(require_editor),
    db: Session = Depends(get_db)
):
    """Delete a report (requires editor or admin role)"""

    repo = ReportRepository(db)
    report = repo.get_by_id(report_id)

    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report not found"
        )

    # Verify tenant access
    if report.tenant_id != current_user.tenant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied"
        )

    repo.delete(report_id)
    return None


@router.post("/bulk-delete", response_model=schemas.BulkDeleteResponse)
async def bulk_delete_reports(
    report_ids: List[str] = Query(..., description="List of report IDs to delete"),
    current_user: User = Depends(require_editor),
    db: Session = Depends(get_db)
):
    """
    Delete multiple reports at once (requires editor or admin role)

    - **report_ids**: List of report IDs to delete
    """
    repo = ReportRepository(db)
    deleted_count = 0
    skipped_count = 0

    for rid in report_ids:
        try:
            report = repo.get_by_id(UUID(rid))
            if report and report.tenant_id == current_user.tenant_id:
                repo.delete(UUID(rid))
                deleted_count += 1
            else:
                skipped_count += 1
        except ValueError:
            skipped_count += 1  # Skip invalid UUIDs

    return schemas.BulkDeleteResponse(
        deleted_count=deleted_count,
        skipped_count=skipped_count
    )


@router.post("/export")
async def export_reports(
    format: str = Query("csv", description="Export format: csv or excel"),
    report_ids: Optional[List[str]] = Query(None, description="Specific report IDs to export"),
    provider: Optional[str] = Query(None, description="Filter by provider"),
    sentiment: Optional[str] = Query(None, description="Filter by sentiment"),
    brand: Optional[str] = Query(None, description="Filter by brand"),
    start_date: Optional[datetime] = Query(None, description="Filter from date"),
    end_date: Optional[datetime] = Query(None, description="Filter to date"),
    search: Optional[str] = Query(None, description="Search in title and summary"),
    current_user: User = Depends(require_viewer),
    db: Session = Depends(get_db)
):
    """
    Export reports to CSV or Excel format

    - **format**: Export format - 'csv' or 'excel' (default: csv)
    - **report_ids**: Optional list of specific report IDs to export
    - **provider**: Filter by provider (if no report_ids specified)
    - **sentiment**: Filter by sentiment
    - **brand**: Filter by brand
    - **start_date**: Filter from date
    - **end_date**: Filter to date
    - **search**: Search in title and summary
    """
    repo = ReportRepository(db)

    # Get reports - either by IDs or by filters
    if report_ids and len(report_ids) > 0:
        # Fetch specific reports by IDs
        reports = []
        for rid in report_ids:
            try:
                report = repo.get_by_id(UUID(rid))
                if report and report.tenant_id == current_user.tenant_id:
                    reports.append(report)
            except ValueError:
                continue  # Skip invalid UUIDs
    else:
        # Use filters to get reports (up to 1000 for export)
        if search:
            reports = repo.search(current_user.tenant_id, search, limit=1000)
        else:
            reports = repo.get_all(
                tenant_id=current_user.tenant_id,
                skip=0,
                limit=1000,
                provider=provider,
                status='completed',
                start_date=start_date,
                end_date=end_date,
                sentiment=sentiment,
                brand=brand
            )

    if not reports:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No reports found to export"
        )

    # Define headers
    headers = ['Title', 'Provider', 'Date', 'Sentiment', 'Topic', 'Brands', 'Reach', 'Source', 'Link', 'Summary']

    if format.lower() == 'excel':
        return _export_to_excel(reports, headers)
    else:
        return _export_to_csv(reports, headers)


def _format_date(dt: datetime) -> str:
    """Format datetime for export"""
    if dt:
        return dt.strftime('%Y-%m-%d %H:%M')
    return ''


def _export_to_csv(reports, headers):
    """Generate CSV file from reports"""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

    # Write header
    writer.writerow(headers)

    # Write data rows
    for report in reports:
        row = [
            report.title or '',
            report.provider or '',
            _format_date(report.timestamp),
            report.sentiment or '',
            report.topic or '',
            ', '.join(report.brands) if report.brands else '',
            str(report.est_reach or 0),
            report.source or '',
            report.link or '',
            report.summary or ''
        ]
        writer.writerow(row)

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=reports.csv"
        }
    )


def _export_to_excel(reports, headers):
    """Generate Excel (.xlsx) file from reports using openpyxl"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"

    # Style for header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Write header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_idx, report in enumerate(reports, 2):
        ws.cell(row=row_idx, column=1, value=report.title or '')
        ws.cell(row=row_idx, column=2, value=report.provider or '')
        ws.cell(row=row_idx, column=3, value=_format_date(report.timestamp))
        ws.cell(row=row_idx, column=4, value=report.sentiment or '')
        ws.cell(row=row_idx, column=5, value=report.topic or '')
        ws.cell(row=row_idx, column=6, value=', '.join(report.brands) if report.brands else '')
        ws.cell(row=row_idx, column=7, value=report.est_reach or 0)
        ws.cell(row=row_idx, column=8, value=report.source or '')
        ws.cell(row=row_idx, column=9, value=report.link or '')
        ws.cell(row=row_idx, column=10, value=report.summary or '')

    # Auto-adjust column widths (approximate)
    column_widths = [40, 15, 20, 12, 15, 30, 12, 30, 50, 60]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Save to bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=reports.xlsx"
        }
    )
