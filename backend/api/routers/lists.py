"""
Lists Router
Handles CRUD operations for lists and list items
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from uuid import UUID
from math import ceil
import csv
import io
import sys
from pathlib import Path

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "src"))

from api.database import get_db
from api import schemas
from api.auth import get_current_user, require_viewer, require_editor, require_admin
from models.user import User
from repositories.list_repository import ListRepository

router = APIRouter()

# Supported list types - add new types here as they become available
SUPPORTED_LIST_TYPES = [
    {"id": "report", "label": "Reports", "description": "Collection of media reports"},
    # Future types (uncomment when implemented):
    # {"id": "contact", "label": "Contacts", "description": "Collection of contacts"},
    # {"id": "editor", "label": "Editors", "description": "Collection of editors"},
]


# ==================== List Types ====================

@router.get("/types/")
async def get_list_types():
    """Get all supported list types"""
    return {"types": SUPPORTED_LIST_TYPES}


# ==================== List CRUD ====================

@router.get("/", response_model=schemas.ListListResponse)
async def list_lists(
    list_type: Optional[str] = Query(None, description="Filter by list type (report, contact, etc)"),
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=100, description="Items per page"),
    current_user: User = Depends(require_viewer),
    db: Session = Depends(get_db)
):
    """
    List all lists for the current user's tenant

    - **list_type**: Filter by list type (optional)
    - **page**: Page number (default: 1)
    - **page_size**: Items per page (default: 50, max: 100)
    """
    repo = ListRepository(db)
    skip = (page - 1) * page_size

    lists = repo.get_all(
        current_user.tenant_id,
        list_type=list_type,
        skip=skip,
        limit=page_size
    )
    total = repo.count(current_user.tenant_id, list_type=list_type)

    return {
        "items": [lst.to_dict() for lst in lists],
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": ceil(total / page_size) if total > 0 else 1
    }


@router.get("/{list_id}", response_model=schemas.ListWithReports)
async def get_list(
    list_id: UUID,
    include_items: bool = Query(True, description="Include items in response"),
    current_user: User = Depends(require_viewer),
    db: Session = Depends(get_db)
):
    """Get a specific list by ID with its items"""
    repo = ListRepository(db)

    if include_items:
        result = repo.get_list_with_reports(list_id)
        if not result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="List not found"
            )
        # Verify tenant access
        list_obj = repo.get_by_id(list_id)
        if list_obj.tenant_id != current_user.tenant_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
        return result
    else:
        list_obj = repo.get_by_id(list_id)
        if not list_obj:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="List not found"
            )
        if list_obj.tenant_id != current_user.tenant_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied"
            )
        return list_obj.to_dict()


@router.post("/", response_model=schemas.ListResponse, status_code=status.HTTP_201_CREATED)
async def create_list(
    list_data: schemas.ListCreate,
    current_user: User = Depends(require_editor),
    db: Session = Depends(get_db)
):
    """
    Create a new list (requires editor role)

    - **name**: Name of the list
    - **list_type**: Type of items in the list (report, contact, editor)
    - **description**: Optional description
    """
    repo = ListRepository(db)

    new_list = repo.create(
        tenant_id=current_user.tenant_id,
        name=list_data.name,
        list_type=list_data.list_type,
        description=list_data.description,
        created_by=current_user.id
    )

    return new_list.to_dict()


@router.put("/{list_id}", response_model=schemas.ListResponse)
@router.patch("/{list_id}", response_model=schemas.ListResponse)
async def update_list(
    list_id: UUID,
    list_update: schemas.ListUpdate,
    current_user: User = Depends(require_editor),
    db: Session = Depends(get_db)
):
    """Update a list (requires editor role)"""
    repo = ListRepository(db)
    list_obj = repo.get_by_id(list_id)

    if not list_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="List not found"
        )

    if list_obj.tenant_id != current_user.tenant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied"
        )

    update_data = list_update.model_dump(exclude_unset=True)
    updated_list = repo.update(list_id, **update_data)

    return updated_list.to_dict()


@router.delete("/{list_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_list(
    list_id: UUID,
    current_user: User = Depends(require_editor),
    db: Session = Depends(get_db)
):
    """Delete a list and all its items (requires editor role)"""
    repo = ListRepository(db)
    list_obj = repo.get_by_id(list_id)

    if not list_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="List not found"
        )

    if list_obj.tenant_id != current_user.tenant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied"
        )

    repo.delete(list_id)
    return None


# ==================== List Item Operations ====================

@router.post("/{list_id}/items/", status_code=status.HTTP_201_CREATED)
async def add_item_to_list(
    list_id: UUID,
    item_data: schemas.ListItemCreate,
    current_user: User = Depends(require_editor),
    db: Session = Depends(get_db)
):
    """Add a single item to a list"""
    repo = ListRepository(db)
    list_obj = repo.get_by_id(list_id)

    if not list_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="List not found"
        )

    if list_obj.tenant_id != current_user.tenant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied"
        )

    item = repo.add_item(list_id, item_data.item_id, current_user.id)

    if not item:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Item already exists in this list"
        )

    return item.to_dict()


@router.post("/{list_id}/items/bulk/", status_code=status.HTTP_201_CREATED)
async def add_items_to_list(
    list_id: UUID,
    items_data: schemas.ListItemBulkAdd,
    current_user: User = Depends(require_editor),
    db: Session = Depends(get_db)
):
    """Add multiple items to a list (skips duplicates)"""
    repo = ListRepository(db)
    list_obj = repo.get_by_id(list_id)

    if not list_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="List not found"
        )

    if list_obj.tenant_id != current_user.tenant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied"
        )

    added_items = repo.add_items(list_id, items_data.item_ids, current_user.id)

    return {
        "added_count": len(added_items),
        "items": [item.to_dict() for item in added_items]
    }


@router.post("/bulk-add/", status_code=status.HTTP_201_CREATED)
async def add_items_to_multiple_lists(
    data: schemas.ListItemMultiListAdd,
    current_user: User = Depends(require_editor),
    db: Session = Depends(get_db)
):
    """Add multiple items to multiple lists at once"""
    repo = ListRepository(db)

    # Verify all lists belong to user's tenant
    for list_id in data.list_ids:
        list_obj = repo.get_by_id(list_id)
        if not list_obj:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"List {list_id} not found"
            )
        if list_obj.tenant_id != current_user.tenant_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Access denied to list {list_id}"
            )

    results = repo.add_items_to_multiple_lists(data.list_ids, data.item_ids, current_user.id)

    return {
        "results": {
            list_id: {
                "added_count": len(items),
                "items": [item.to_dict() for item in items]
            }
            for list_id, items in results.items()
        }
    }


@router.delete("/{list_id}/items/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
async def remove_item_from_list(
    list_id: UUID,
    item_id: UUID,
    current_user: User = Depends(require_editor),
    db: Session = Depends(get_db)
):
    """Remove a single item from a list"""
    repo = ListRepository(db)
    list_obj = repo.get_by_id(list_id)

    if not list_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="List not found"
        )

    if list_obj.tenant_id != current_user.tenant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied"
        )

    removed = repo.remove_item(list_id, item_id)

    if not removed:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Item not found in this list"
        )

    return None


@router.delete("/{list_id}/items/bulk/")
async def remove_items_from_list(
    list_id: UUID,
    items_data: schemas.ListItemBulkAdd,
    current_user: User = Depends(require_editor),
    db: Session = Depends(get_db)
):
    """Remove multiple items from a list"""
    repo = ListRepository(db)
    list_obj = repo.get_by_id(list_id)

    if not list_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="List not found"
        )

    if list_obj.tenant_id != current_user.tenant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied"
        )

    removed_count = repo.remove_items(list_id, items_data.item_ids)

    return {"removed_count": removed_count}


# ==================== Query Helpers ====================

@router.get("/containing/{item_id}")
async def get_lists_containing_item(
    item_id: UUID,
    current_user: User = Depends(require_viewer),
    db: Session = Depends(get_db)
):
    """Get all lists that contain a specific item"""
    repo = ListRepository(db)
    lists = repo.get_lists_containing_item(current_user.tenant_id, item_id)

    return [lst.to_dict() for lst in lists]


# ==================== Export ====================

@router.post("/{list_id}/export/")
async def export_list(
    list_id: UUID,
    format: str = Query("csv", description="Export format: csv or excel"),
    current_user: User = Depends(require_viewer),
    db: Session = Depends(get_db)
):
    """Export a list's contents to CSV or Excel"""
    repo = ListRepository(db)
    list_obj = repo.get_by_id(list_id)

    if not list_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="List not found"
        )

    if list_obj.tenant_id != current_user.tenant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied"
        )

    # Get reports in the list
    reports = repo.get_reports_in_list(list_id, limit=10000)

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow([
            "Title", "Source", "Provider", "Brands", "Sentiment",
            "Topic", "Reach", "Link", "Date", "Summary"
        ])

        # Data rows
        for report in reports:
            writer.writerow([
                report.title,
                report.source,
                report.provider,
                ", ".join(report.brands or []),
                report.sentiment,
                report.topic,
                report.est_reach,
                report.link,
                report.timestamp.isoformat() if report.timestamp else "",
                report.summary or ""
            ])

        output.seek(0)
        filename = f"{list_obj.name.replace(' ', '_')}_export.csv"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    elif format == "excel":
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_501_NOT_IMPLEMENTED,
                detail="Excel export not available (openpyxl not installed)"
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reports"

        # Header
        headers = [
            "Title", "Source", "Provider", "Brands", "Sentiment",
            "Topic", "Reach", "Link", "Date", "Summary"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Data rows
        for row, report in enumerate(reports, 2):
            ws.cell(row=row, column=1, value=report.title)
            ws.cell(row=row, column=2, value=report.source)
            ws.cell(row=row, column=3, value=report.provider)
            ws.cell(row=row, column=4, value=", ".join(report.brands or []))
            ws.cell(row=row, column=5, value=report.sentiment)
            ws.cell(row=row, column=6, value=report.topic)
            ws.cell(row=row, column=7, value=report.est_reach)
            ws.cell(row=row, column=8, value=report.link)
            ws.cell(row=row, column=9, value=report.timestamp.isoformat() if report.timestamp else "")
            ws.cell(row=row, column=10, value=report.summary or "")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{list_obj.name.replace(' ', '_')}_export.xlsx"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid format. Use 'csv' or 'excel'"
        )
